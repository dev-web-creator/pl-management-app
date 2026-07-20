# -*- coding: utf-8 -*-
# 人生設計シート(xlsx) → pl_app インポートSQL生成（ADR-045）
# 方針:
#  - 全行に決定的 client_key（imp:種別:期間:項目）→ 何度流しても冪等
#  - 取引は支払い脚なし（現在の残高計算に影響させない）
#  - カットオフ: 2026-07-01（以降はアプリが正。targets のみ未来月を投入）
import openpyxl, os, sys, io
from datetime import date, timedelta

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
XLSX = os.path.expandvars(r'%TEMP%\claude\sheet.xlsx')
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'import.sql')
CUTOFF = '2026-07-01'

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

def serial2date(n):
    return date(1899, 12, 30) + timedelta(days=int(n))

def month_end(y, m):
    nm = date(y + (m == 12), (m % 12) + 1, 1)
    return nm - timedelta(days=1)

def num(v):
    if v is None or v in ('', '-', '/'): return None
    try: return round(float(v))
    except (TypeError, ValueError): return None

def esc(s): return str(s).replace("'", "''")

# ---------------- カテゴリマッピング ----------------
DAILY_MAP = {  # 日次シート列 → アプリのカテゴリ名
    '朝飯': '朝飯', '昼飯': '昼飯', '晩飯': '晩飯', '交際費': '交際費',
    'プレゼント': 'プレゼント・奢り', 'スーパー': 'スーパー・まとめ買い',
    '体験': '1人体験', '結婚式': '結婚式', '旅費・交通費': '旅費・交通費',
    '物品購入費': '物品購入費', 'その他諸経費': 'その他諸経費',
    '趣味・嗜好品(TBC等)': '嗜好品(TBC)',
}
def fixed_map(label):  # 月次シートの固定費行ラベル → カテゴリ名
    l = label
    if '家賃' in l: return '家賃'
    if 'ジム' in l: return 'ジム'
    if '生命保険' in l: return '生命保険'
    if '携帯' in l or 'YouTubeプレミアム' == l: return '携帯+YouTube'
    if '定期券' in l: return '定期券代'
    if 'iCloud' in l: return 'iCloud拡張'
    if 'BASE FOOD' in l: return 'BASE FOOD/ナッシュ'
    if 'Kindle' in l: return 'Kindle'
    if 'Netflix' in l: return 'Netflix'
    if 'Abema' in l: return 'Abemaプレミアム'
    if '脱毛' in l: return '脱毛(プラム)'
    if 'ChatGPT' in l.replace('chat gpt', 'ChatGPT').replace('　', ''): return 'ChatGPTプラス'
    if 'しがく' in l: return 'しがく会費'
    if 'Amazon Prime' in l: return 'Amazon Prime'
    if 'テレ東' in l: return 'テレ東BIZ'
    return None
INCOME_MAP = [  # (含む文字列, カテゴリ名, 取引type)
    ('給与収入', '給与収入(手取り)', 'income'),
    ('経費精算', '経費精算', 'income'),
    ('家族収入', '家族収入', 'income'),
    ('その他収益', '副業・その他収益', 'income'),
    ('その他収入', '副業・その他収益', 'income'),
    ('ポイント収入', 'ポイント収入', 'income'),
    ('借入金', '借入金', 'income'),
]
FY1_VAR_MAP = {  # FY-1 変動費行 → アプリのカテゴリ名（原ラベルはmemoに保存）
    '食費（1人）': '(取込)食費一括',
    '①交際費（誰かと飲み・ごはん）': '交際費',
    '③温泉費用（館内での飲食含む）': '1人体験',
    'タバコ': '嗜好品(TBC)',
    'スーパー購入費': 'スーパー・まとめ買い',
    '⑩映画・ドラマ（netflixの固定費以外でかかった費用）': '1人体験',
    '⑨ライブ・スポーツ観戦など（ex,甲子園）': '1人体験',
    '⑦香水・フレグランス': '物品購入費',
    '②旅行費（名古屋・有馬・富士急など）': '旅費・交通費',
    'プレゼント・おごりなど': 'プレゼント・奢り',
    '結婚式': '結婚式',
    '1人体験(運動,散髪,グリーン車,1人映画鑑賞)': '1人体験',
    '⑧本・セミナー・イベントなど': '1人体験',
    '④体験(グリーン車やホテル,個サルや運動など)': '1人体験',
    '⑤散髪': '1人体験',
    '⑥物品購入費(衣類・家具・髭剃りなど生活用品全般)': '物品購入費',
}
FY1_VAR_CONTAINS = [  # 上記の完全一致で拾えない場合の部分一致
    ('食費（1人）', '(取込)食費一括'), ('①交際費', '交際費'), ('温泉', '1人体験'),
    ('タバコ', '嗜好品(TBC)'), ('スーパー', 'スーパー・まとめ買い'), ('映画・ドラマ', '1人体験'),
    ('ライブ・スポーツ', '1人体験'), ('香水', '物品購入費'), ('②旅行費', '旅費・交通費'),
    ('プレゼント', 'プレゼント・奢り'), ('結婚式', '結婚式'), ('1人体験', '1人体験'),
    ('本・セミナー', '1人体験'), ('④体験', '1人体験'), ('散髪', '1人体験'),
    ('旅費・交通費', '旅費・交通費'), ('物品購入費', '物品購入費'), ('その他諸経費', 'その他諸経費'),
]
NEW_CATEGORIES = [  # (name, pl_type, parent名 or None, is_input_allowed)
    ('脱毛(プラム)', 'fixed_cost', '固定費', True),
    ('ChatGPTプラス', 'fixed_cost', '固定費', True),
    ('しがく会費', 'fixed_cost', '固定費', True),
    ('Amazon Prime', 'fixed_cost', '固定費', True),
    ('テレ東BIZ', 'fixed_cost', '固定費', True),
    ('(取込)食費一括', 'variable_cost', '食費(1人)', False),
]
DEDUCTIONS = ['住民税', '所得税', '健康保険料', '雇用保険料', '厚生年金保険料',
              '子ども・子育て支援金', '定額減税', '年末調整精算額', 'ふるさと納税']
def ded_map(label):
    for d in DEDUCTIONS:
        if label.startswith(d) or d in label:
            if d == '定額減税': return '定額減税・確定申告還付'
            if d == 'ふるさと納税': return 'ふるさと納税返金額'
            return d
    return None
WALLET_CONTAINS = [  # 資産管理の行ラベル → wallet名（先勝ち）
    ('PayPayポイント', 'PayPayポイント(自動運用)'), ('PayPay残高', 'PayPay残高'),
    ('みずほ', 'みずほ銀行'), ('UFJ', '三菱UFJ銀行'), ('三井住友', '三井住友銀行'),
    ('楽天銀行', '楽天銀行'), ('PASMO', 'PASMO'), ('ANAPay', 'ANA Pay'), ('ANA Pay', 'ANA Pay'),
    ('JALPay', 'JAL Pay'), ('JAL Pay', 'JAL Pay'), ('Vポイント', 'VポイントPay'),
    ('メルペイ', 'メルペイ'), ('Ponta', 'Pontaポイント'), ('dポイント', 'dポイント'),
    ('ICOCA', 'ICOCA'),
    ('イーサリアム', 'bitFlyer ETH'), ('ビットコイン', 'bitFlyer BTC'), ('エックスアールピー', 'bitFlyer XRP'),
]
SKIP_WALLET_ROWS = ('銀行預金', 'キャッシュレス残高', '暗号通貨', '買い注文', '為替', '価格', 'USD',
                    '合計', '純資産', '目標', '達成', '証券', '株式', 'サンリオ', 'iDeCo', '投資')

sql = []
warn = []
stats = {}
def add(kind, n=1): stats[kind] = stats.get(kind, 0) + n

sql.append('-- 人生設計シート → pl_app インポート（自動生成・冪等）')
sql.append('BEGIN;')
# 新カテゴリ
for name, pl, parent, allowed in NEW_CATEGORIES:
    p = f"(SELECT id FROM categories WHERE user_id=1 AND name='{esc(parent)}')" if parent else 'NULL'
    sql.append(
        f"INSERT INTO categories (user_id, parent_id, name, pl_type, is_input_allowed, display_order)\n"
        f"SELECT 1, {p}, '{esc(name)}', '{pl}', {str(allowed).lower()}, 900\n"
        f"WHERE NOT EXISTS (SELECT 1 FROM categories WHERE user_id=1 AND name='{esc(name)}');")
# XRPウォレット（無ければ）
sql.append("INSERT INTO wallets (user_id, name, type, display_order)\n"
           "SELECT 1, 'bitFlyer XRP', 'crypto', 102\n"
           "WHERE NOT EXISTS (SELECT 1 FROM wallets WHERE user_id=1 AND name='bitFlyer XRP');")
# 過去データ用ウォレットは作らない（脚なしインポート）

def tx(dt, cat, amount, memo, ckey, txtype='expense'):
    if amount is None or amount == 0: return
    if str(dt) >= CUTOFF: return
    sql.append(
        f"INSERT INTO transactions (user_id, category_id, type, amount, accrual_date, memo, client_key)\n"
        f"SELECT 1, c.id, '{txtype}', {abs(amount)}, '{dt}', '{esc(memo)}', '{esc(ckey)}'\n"
        f"FROM categories c WHERE c.user_id=1 AND c.name='{esc(cat)}'\n"
        f"ON CONFLICT (user_id, client_key) WHERE client_key IS NOT NULL DO NOTHING;")
    add('tx:' + ('income' if txtype == 'income' else 'expense'))

# ---------------- 1) 日次出費（FY-2 / FY-3） ----------------
daily_sum = {}  # (tag, 'YYYY-MM', appカテゴリ) → 合計（月次タブとの突合・調整用）
for sheet, tag in [('FY-2(日次)出費', 'd2'), ('FY-3(日次)出費', 'd3')]:
    ws = wb[sheet]
    header = None
    for row in ws.iter_rows(min_row=4, values_only=True):
        if header is None:
            header = row  # 行4: 【日付】,合計,朝飯...
            cols = {}
            for i, h in enumerate(header):
                if h in DAILY_MAP: cols[i] = DAILY_MAP[h]
            continue
        d = row[1]
        if d is None or not isinstance(d, (int, float)): continue
        dt = serial2date(d)
        for i, cat in cols.items():
            v = num(row[i]) if i < len(row) else None
            if v:
                tx(dt, cat, v, f'シート取込({sheet.split("(")[0]}日次)', f'imp:{tag}:{dt}:{header[i][:12]}')
                if str(dt) < CUTOFF:
                    k = (tag, f'{dt.year}-{dt.month:02d}', cat)
                    daily_sum[k] = daily_sum.get(k, 0) + v

# ---------------- 2) 月次支出：固定費＋支払い利息（FY-2 / FY-3） ----------------
for sheet, tag in [('FY-2(月次)支出', 'f2'), ('FY-3(月次)支出 ', 'f3')]:
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    marks, heads = rows[0], rows[1]
    months = {}  # 列index → (y,m)
    for i, h in enumerate(heads):
        if isinstance(h, str) and h.endswith('月') and '年' in h and str(marks[i]).startswith('※確定'):
            y, m = h.replace('月', '').split('年')
            months[i] = (int(y), int(m))
    section = 'fixed'
    for r in rows[2:]:
        b = str(r[1]) if r[1] else ''
        c = str(r[2]) if r[2] else ''
        if '変動費' in b: section = 'var'
        if '支払い利息' in b: section = 'interest'
        if '税金' in b: section = 'tax'
        if not c or '合計' in c or '比率' in c: continue
        if section == 'fixed':
            cat = fixed_map(c)
            if cat is None:
                warn.append(f'{sheet}: 固定費行 未マップ: {c}')
                continue
            for i, (y, m) in months.items():
                v = num(r[i])
                if v: tx(date(y, m, 1), cat, v, f'シート取込(固定費): {c[:40]}', f'imp:{tag}:{y}-{m:02d}:{cat[:12]}')
        elif section == 'var':
            # 月次タブの変動費リーフ＝確定値。日次の集計との正の差分を「月次調整」として取込
            # （引越し費など月次タブに直接入力された分の救済。負の差分は警告のみ）
            MVAR = [('食費（1人', None), ('交際費', '交際費'), ('プレゼント', 'プレゼント・奢り'),
                    ('スーパー', 'スーパー・まとめ買い'), ('1人体験', '1人体験'), ('結婚式', '結婚式'),
                    ('旅費・交通費', '旅費・交通費'), ('物品購入費', '物品購入費'),
                    ('その他諸経費', 'その他諸経費'), ('嗜好品', '嗜好品(TBC)')]
            mcat = None
            for key, cat in MVAR:
                if c.startswith(key): mcat = key if cat is None else cat; break
            if mcat is None: continue
            dtag = 'd2' if tag == 'f2' else 'd3'
            atag = 'a2' if tag == 'f2' else 'a3'
            for i, (y, m) in months.items():
                v = num(r[i])
                if v is None: continue
                mo = f'{y}-{m:02d}'
                if mcat == '食費（1人':
                    daily = sum(daily_sum.get((dtag, mo, x), 0) for x in ('朝飯', '昼飯', '晩飯'))
                    target_cat = '(取込)食費一括'
                else:
                    daily = daily_sum.get((dtag, mo, mcat), 0)
                    target_cat = mcat
                diff = v - daily
                if diff > 0:
                    tx(month_end(y, m), target_cat, diff, f'シート取込(月次調整): {c[:36]}', f'imp:{atag}:{mo}:{target_cat[:12]}')
                elif diff < 0:
                    warn.append(f'{sheet} {mo} {c[:20]}: 月次({v}) < 日次集計({daily}) 差{diff} → 日次を正とし調整なし')
        elif section == 'interest':
            if '返済利子' in c:
                for i, (y, m) in months.items():
                    v = num(r[i])
                    if v: tx(date(y, m, 1), '支払い利息', v, 'シート取込: 返済利子', f'imp:{tag}:{y}-{m:02d}:利子')
            elif '元金返済' in c:
                for i, (y, m) in months.items():
                    v = num(r[i])
                    if v: tx(date(y, m, 1), '元金返済', v, 'シート取込: 元金返済', f'imp:{tag}:{y}-{m:02d}:元金')

# ---------------- 3) 月次収入＋給与明細（FY-2 / FY-3） ----------------
for sheet, tag in [('FY-2(月次)収入', 'i2'), ('FY-3(月次)収入', 'i3')]:
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    marks, heads = rows[0], rows[1]
    months = {}
    for i, h in enumerate(heads):
        if isinstance(h, str) and h.endswith('月') and '年' in h and str(marks[i]).startswith('※確定'):
            y, m = h.replace('月', '').split('年')
            months[i] = (int(y), int(m))
    section = 'income'
    payslip = {k: {'gross': None, 'ded': [], 'allow': [], 'hours': None, 'ot': None} for k in months.values()}
    for r in rows[2:]:
        b = str(r[1]) if r[1] else ''
        c = str(r[2]) if r[2] else ''
        if '控除額内訳' in b: section = 'ded'
        if '給与内訳' in b: section = 'allow'
        if not c or c.startswith('月次合計') or '比率' in c: continue
        if section == 'income':
            if '総支給額' in c:
                for i, (y, m) in months.items():
                    v = num(r[i])
                    if v: payslip[(y, m)]['gross'] = v
                continue
            for key, cat, tt in INCOME_MAP:
                if key in c:
                    for i, (y, m) in months.items():
                        v = num(r[i])
                        if v: tx(date(y, m, 25), cat, v, f'シート取込(収入): {c[:40]}', f'imp:{tag}:{y}-{m:02d}:{cat[:10]}', 'income')
                    break
            else:
                warn.append(f'{sheet}: 収入行 未マップ: {c}')
        elif section == 'ded':
            d = ded_map(c)
            if d:
                for i, (y, m) in months.items():
                    v = num(r[i])
                    if v: payslip[(y, m)]['ded'].append((d, v))
        elif section == 'allow' and tag == 'i3':  # FY-2の給与内訳は破損しているためFY-3のみ
            if '総労働時間' == c:
                for i, (y, m) in months.items(): payslip[(y, m)]['hours'] = r[i]
            elif '時間外労働時間' == c:
                for i, (y, m) in months.items(): payslip[(y, m)]['ot'] = r[i]
            elif '時給' in c or '残業代' in c: continue
            else:
                for i, (y, m) in months.items():
                    v = num(r[i])
                    if v: payslip[(y, m)]['allow'].append((c[:40], v))
    # 給与明細SQL
    for (y, m), p in payslip.items():
        if p['gross'] is None and not p['ded']: continue
        allows = p['allow'] if p['allow'] else ([('総支給(シート取込)', p['gross'])] if p['gross'] else [])
        items = [f"('allowance','{esc(n)}',{v})" for n, v in allows] + [f"('deduction','{esc(n)}',{v})" for n, v in p['ded']]
        if not items: continue
        h = 'NULL' if p['hours'] in (None, '', 0) else round(float(p['hours']), 1)
        o = 'NULL' if p['ot'] in (None, '', 0) else round(float(p['ot']), 1)
        sql.append(
            f"WITH ins AS (\n"
            f"  INSERT INTO payslips (user_id, period, total_work_hours, overtime_hours, is_confirmed, source)\n"
            f"  VALUES (1, '{y}-{m:02d}-01', {h}, {o}, true, 'manual')\n"
            f"  ON CONFLICT (user_id, period) DO NOTHING RETURNING id)\n"
            f"INSERT INTO payslip_items (payslip_id, item_type, name, amount)\n"
            f"SELECT ins.id, v.t, v.n, v.a FROM ins, (VALUES {','.join(items)}) AS v(t,n,a);")
        add('payslip')

# ---------------- 4) FY-1 月次（支出予実・収入予実） ----------------
ws = wb['FY-1(月次)支出予実']
rows = list(ws.iter_rows(values_only=True))
months1 = {}
for i, h in enumerate(rows[1]):
    if isinstance(h, (int, float)) and h > 40000 and str(rows[0][i]).startswith('※確定'):
        d = serial2date(h)
        if (d.year, d.month) == (2025, 4): continue  # FY-2タブと重複する境界月はFY-2を正とする
        months1[i] = (d.year, d.month)
section = 'fixed'
for r in rows[2:]:
    b = str(r[1]) if r[1] else ''
    if not b or '合計' in b or '比率' in b or b == '☑':
        if '月次合計' in b and section == 'fixed' and rows.index(r) > 18: section = 'var'
        continue
    if '+' in b: continue  # 合成行（①交際費+②旅行費 / ⑥物品購入+⑦香水+⑧本 等）＝二重計上防止
    if '返済利子' in b:
        for i, (y, m) in months1.items():
            v = num(r[i])
            if v: tx(date(y, m, 1), '支払い利息', v, 'シート取込(FY-1): 返済利子', f'imp:f1:{y}-{m:02d}:利子')
        continue
    if '元金返済' in b:
        for i, (y, m) in months1.items():
            v = num(r[i])
            if v: tx(date(y, m, 1), '元金返済', v, 'シート取込(FY-1): 元金返済', f'imp:f1:{y}-{m:02d}:元金')
        continue
    if '給与天引き' in b or b == '合算費用': continue  # 計上なし注記行・総計行
    if section == 'fixed' or fixed_map(b):
        cat = fixed_map(b)
        if cat:
            for i, (y, m) in months1.items():
                v = num(r[i])
                # キーは行ラベル基準（同一カテゴリに2行マップされても衝突しない）
                if v: tx(date(y, m, 1), cat, v, f'シート取込(FY-1固定費): {b[:40]}', f'imp:f1:{y}-{m:02d}:{b[:12]}')
            continue
    matched = FY1_VAR_MAP.get(b)
    if not matched:
        for key, cat in FY1_VAR_CONTAINS:
            if key in b: matched = cat; break
    if matched:
        for i, (y, m) in months1.items():
            v = num(r[i])
            if v: tx(date(y, m, 1), matched, v, f'シート取込(FY-1): {b[:40]}', f'imp:v1:{y}-{m:02d}:{b[:12]}')
    else:
        warn.append(f'FY-1支出: 未マップ: {b}')

ws = wb['FY-1(月次)収入予実']
rows = list(ws.iter_rows(values_only=True))
months1i = {}
for i, h in enumerate(rows[1]):
    if isinstance(h, (int, float)) and h > 40000 and str(rows[0][i]).startswith('※確定'):
        d = serial2date(h)
        if (d.year, d.month) == (2025, 4): continue  # FY-2タブと重複する境界月はFY-2を正とする
        months1i[i] = (d.year, d.month)
section = 'income'
payslip1 = {k: {'gross': None, 'ded': []} for k in months1i.values()}
for r in rows[2:]:
    e = str(r[4]) if len(r) > 4 and r[4] else ''
    b = str(r[1]) if r[1] else ''
    if '控除額内訳' in b: section = 'ded'
    if not e or e.startswith('月次合計'): continue
    if section == 'income':
        if e.startswith('総支給額'):
            for i, (y, m) in months1i.items():
                v = num(r[i])
                if v: payslip1[(y, m)]['gross'] = v
            continue
        if 'TikTok' in e:
            for i, (y, m) in months1i.items():
                v = num(r[i])
                if v: tx(date(y, m, 25), 'ポイント収入', v, 'シート取込(FY-1): TikTokポイント収入', f'imp:i1:{y}-{m:02d}:ポイント', 'income')
            continue
        for key, cat, tt in INCOME_MAP:
            if key in e:
                for i, (y, m) in months1i.items():
                    v = num(r[i])
                    if v: tx(date(y, m, 25), cat, v, f'シート取込(FY-1): {e[:40]}', f'imp:i1:{y}-{m:02d}:{cat[:10]}', 'income')
                break
        else:
            warn.append(f'FY-1収入: 未マップ: {e}')
    else:
        d = ded_map(e)
        if d:
            for i, (y, m) in months1i.items():
                v = num(r[i])
                if v: payslip1[(y, m)]['ded'].append((d, v))
for (y, m), p in payslip1.items():
    if p['gross'] is None: continue
    items = [f"('allowance','総支給(シート取込)',{p['gross']})"] + [f"('deduction','{esc(n)}',{v})" for n, v in p['ded']]
    sql.append(
        f"WITH ins AS (\n"
        f"  INSERT INTO payslips (user_id, period, is_confirmed, source) VALUES (1, '{y}-{m:02d}-01', true, 'manual')\n"
        f"  ON CONFLICT (user_id, period) DO NOTHING RETURNING id)\n"
        f"INSERT INTO payslip_items (payslip_id, item_type, name, amount)\n"
        f"SELECT ins.id, v.t, v.n, v.a FROM ins, (VALUES {','.join(items)}) AS v(t,n,a);")
    add('payslip')

# ---------------- 5) 資産管理 → balance_snapshots（確定月のみ） ----------------
for sheet in ['(FY-2)資産管理', '(FY-3)資産管理']:
    ws = wb[sheet]
    rows = list(ws.iter_rows(max_row=60, values_only=True))
    marks = rows[1]
    months = {}
    for i, h in enumerate(rows[2]):
        if isinstance(h, str) and h.endswith('月') and '年' in h and str(marks[i]).startswith('※確定'):
            y, m = h.replace('月', '').split('年')
            months[i] = (int(y), int(m))
    for r in rows[3:]:
        label = str(r[1]) if r[1] else ''
        if not label or any(s in label for s in SKIP_WALLET_ROWS): continue
        wname = None
        for key, w in WALLET_CONTAINS:
            if key in label: wname = w; break
        if wname is None:
            if any(num(r[i]) for i in months): warn.append(f'{sheet}: ウォレット未マップ: {label}')
            continue
        for i, (y, m) in months.items():
            v = num(r[i])
            if v is None: continue
            if wname.startswith('bitFlyer') and v == 0: continue
            d = month_end(y, m)
            if str(d) >= CUTOFF: continue
            sql.append(
                f"INSERT INTO balance_snapshots (user_id, wallet_id, as_of_date, actual_balance)\n"
                f"SELECT 1, w.id, '{d}', {v} FROM wallets w WHERE w.user_id=1 AND w.name='{esc(wname)}'\n"
                f"ON CONFLICT (wallet_id, as_of_date) DO NOTHING;")
            add('snapshot')

# ---------------- 6) FY-3 未確定月 → targets（income/expense） ----------------
for sheet, metric, rowlabel in [('FY-3(月次)支出 ', 'expense', '合計支出額'), ('FY-3(月次)収入', 'income', '月次合計')]:
    ws = wb[sheet]
    rows = list(ws.iter_rows(max_row=4, values_only=True))
    marks, heads, total = rows[0], rows[1], rows[2]
    for i, h in enumerate(heads):
        if isinstance(h, str) and h.endswith('月') and '年' in h and '未確定' in str(marks[i]):
            y, m = h.replace('月', '').split('年')
            v = num(total[i])
            if v:
                sql.append(
                    f"INSERT INTO targets (user_id, period, metric, amount) VALUES (1, '{int(y)}-{int(m):02d}-01', '{metric}', {v})\n"
                    f"ON CONFLICT (user_id, period, metric) DO NOTHING;")
                add('target')

sql.append('COMMIT;')
with open(OUT, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql) + '\n')

print('生成完了:', OUT)
print('統計:', stats)
print('警告:')
for w in warn: print(' -', w)
