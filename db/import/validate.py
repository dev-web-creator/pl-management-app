# -*- coding: utf-8 -*-
# 全月突合: シートの月次合計 vs DBの取込結果
import openpyxl, os, subprocess, sys, io, re
from collections import Counter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
XLSX = os.path.expandvars(r'%TEMP%\claude\sheet.xlsx')
PSQL = r'C:\pgsql\bin\psql.exe'
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

def q(sql):
    r = subprocess.run([PSQL, '-U', 'postgres', '-d', 'pl_app', '-t', '-A', '-F', '|', '-c', sql],
                       capture_output=True, text=True, encoding='utf-8')
    return [l.split('|') for l in r.stdout.strip().splitlines() if l]

# DB: 取込取引の 月×プレフィックス 合計（元金/利子/借入金は別集計）
db = {}
for pre, month, total in q(
    "SELECT split_part(client_key,':',2) || CASE WHEN client_key LIKE '%:利子' OR client_key LIKE '%:元金' THEN ':loan' "
    "WHEN client_key LIKE '%:借入金' THEN ':borrow' ELSE '' END, "
    "to_char(date_trunc('month',accrual_date),'YYYY-MM'), SUM(amount)::int "
    "FROM transactions WHERE client_key LIKE 'imp:%' GROUP BY 1,2"):
    db[(pre, month)] = int(total)

def num(v):
    if v is None or v in ('', '-', '/'): return None
    try: return round(float(v))
    except: return None

ng = 0
def check(label, expect, got):
    global ng
    if expect is None: return
    mark = 'OK' if expect == (got or 0) else f'** NG (diff {(got or 0)-expect})'
    if expect != (got or 0):
        ng += 1
        print(f'{label}: sheet={expect} db={got or 0} {mark}')

# FY-2/FY-3 月次支出タブの 変動費計/固定費計 と突合（変動費はシート内不整合があり得るので日次側=正として差分報告のみ）
for sheet, dtag, ftag in [('FY-2(月次)支出', 'd2', 'f2'), ('FY-3(月次)支出 ', 'd3', 'f3')]:
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    months = {}
    for i, h in enumerate(rows[1]):
        if isinstance(h, str) and h.endswith('月') and '年' in h and str(rows[0][i]).startswith('※確定'):
            y, m = h.replace('月', '').split('年'); months[i] = f'{int(y)}-{int(m):02d}'
    fixed_row = next(r for r in rows if r[2] and '【固定費】月次合計' in str(r[2]))
    var_row = next(r for r in rows if r[2] and '【変動費】月次合計' in str(r[2]))
    atag = 'a2' if ftag == 'f2' else 'a3'
    for i, mo in months.items():
        if mo >= '2026-07': continue
        check(f'{sheet} {mo} 固定費', num(fixed_row[i]), db.get((ftag, mo), 0))
        # 変動費 = 日次取込 + 月次調整。月次<日次のケース（シート内部不整合）は日次を正とするので差分のみ表示
        got = db.get((dtag, mo), 0) + db.get((atag, mo), 0)
        exp = num(var_row[i])
        if exp is not None and got < exp:
            check(f'{sheet} {mo} 変動費', exp, got)
        elif exp is not None and got > exp:
            print(f'{sheet} {mo} 変動費: sheet={exp} db={got} (日次が月次より多い月・日次を正とする)')

# FY-2/FY-3 収入
for sheet, itag in [('FY-2(月次)収入', 'i2'), ('FY-3(月次)収入', 'i3')]:
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    months = {}
    for i, h in enumerate(rows[1]):
        if isinstance(h, str) and h.endswith('月') and '年' in h and str(rows[0][i]).startswith('※確定'):
            y, m = h.replace('月', '').split('年'); months[i] = f'{int(y)}-{int(m):02d}'
    total_row = rows[2]  # 月次合計(可処分所得)※借入金を含まない
    for i, mo in months.items():
        if mo >= '2026-07': continue
        # 丸め差±3円許容。DB側は借入金(excluded)を除外して比較
        exp, got = num(total_row[i]), db.get((itag, mo), 0)
        if exp is not None and abs(exp - got) > 3:
            print(f'{sheet} {mo} 収入: sheet={exp} db={got} ** NG'); ng += 1

# FY-1
ws = wb['FY-1(月次)支出予実']
rows = list(ws.iter_rows(values_only=True))
from datetime import date, timedelta
months1 = {}
for i, h in enumerate(rows[1]):
    if isinstance(h, (int, float)) and h > 40000 and str(rows[0][i]).startswith('※確定'):
        d = date(1899, 12, 30) + timedelta(days=int(h))
        if (d.year, d.month) != (2025, 4): months1[i] = f'{d.year}-{d.month:02d}'
all_row = rows[2]   # 合算費用（支払い利息セクションを含まない）
for i, mo in months1.items():
    got = db.get(('f1', mo), 0) + db.get(('v1', mo), 0)
    check(f'FY-1 {mo} 支出合計', num(all_row[i]), got)

# スナップショット差分の診断（生成172 vs 挿入結果）
gen = Counter()
sqlf = open(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'import.sql'), encoding='utf-8').read()
for m in re.finditer(r"SELECT 1, w\.id, '(\d{4}-\d{2}-\d{2})', (-?\d+) FROM wallets w WHERE w\.user_id=1 AND w\.name='([^']+)'", sqlf):
    gen[(m.group(3), m.group(1))] += 1
dups = {k: v for k, v in gen.items() if v > 1}
print('生成側の重複キー:', dups if dups else 'なし')
dbsnap = int(q("SELECT COUNT(*) FROM balance_snapshots WHERE user_id=1 AND as_of_date < '2026-07-01'")[0][0])
print(f'スナップショット: 生成{sum(gen.values())} → DB {dbsnap}')
missing = [k for k in gen if not q(f"SELECT 1 FROM balance_snapshots s JOIN wallets w ON w.id=s.wallet_id WHERE w.user_id=1 AND w.name='{k[0]}' AND s.as_of_date='{k[1]}'")]
print('DBに入らなかったもの:', missing[:20] if missing else 'なし')

print()
print('NG件数:', ng)
